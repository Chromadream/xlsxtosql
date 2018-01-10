def importfromexcel(filename,worksheet,rowsize,colsize):
    from openpyxl import load_workbook
    export = []
    wb = load_workbook(filename = filename)
    ws = wb[worksheet]
    for x in range(1,rowsize+1):
        li = []
        for y in range(1,colsize+1):
            li.append(ws.cell(row=x,column=y).value);
        export.append(li)
    return export
